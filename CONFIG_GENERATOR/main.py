import pandas as pd
from jinja2 import Template
from functions import (check_input_file_existence,validate_field_names,check_mandatory_fields,fix_trailing_spaces,check_duplicate_urls)
from filebeat_yaml_template import filebeat_yaml_template
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def update_master_excel(master_excel_file, df, sheet_name, header_start_row=1):
    """
    Update or create a specific sheet in the master Excel file with data from a DataFrame.
    If sheet exists, data will be appended. If not, a new sheet will be created.
    
    """
    # Ensure the directory exists
    os.makedirs(os.path.dirname(master_excel_file), exist_ok=True)
    
    try:
        # Try to load the workbook
        book = load_workbook(master_excel_file)
    except FileNotFoundError:
        # If the workbook does not exist, create it
        book = pd.ExcelWriter(master_excel_file, engine='openpyxl').book

    if sheet_name in book.sheetnames:
        # If the sheet exists, load it
        sheet = book[sheet_name]
        # Find the last row in the sheet that contains data
        last_row = sheet.max_row
        # Append data from DataFrame, skipping the header
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=last_row + 1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
    else:
        # If the sheet does not exist, create it
        sheet = book.create_sheet(sheet_name)
        # Adjust starting row for headers and data based on header_start_row
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=header_start_row):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    book.save(master_excel_file)
    book.close()

def filebeat_config(input_excel, master_excel):
    filebeat_df = pd.read_excel(input_excel, sheet_name='filebeat_input' ,engine='openpyxl', index_col=None, header=2 )
    print(filebeat_df.columns)
    if not filebeat_df.empty:
        validate_field_names(filebeat_df)
        check_mandatory_fields(filebeat_df)
        filebeat_df = fix_trailing_spaces(filebeat_df)
        check_duplicate_urls(filebeat_df,master_excel)
        

        template_filebeat = Template(filebeat_yaml_template)
        yaml_data_filebeat = template_filebeat.render(data=filebeat_df)
        yaml_filebeat = 'filebeat_config.yaml'
        with open(yaml_filebeat, 'w') as f:
            f.write(yaml_data_filebeat)

        print(f"YAML data has been successfully written to {yaml_filebeat}.")

        now = datetime.now()
        filebeat_df['date'] = now.strftime('%d-%m-%y')
        filebeat_df['time'] = now.strftime('%H:%M:%S')
        update_master_excel(master_excel, filebeat_df, 'filebeat_master')


def main():
    input_excel= r'CONFIG_GENERATOR\filebeat_input.xlsx'
    
    master_excel = r'CONFIG_GENERATOR\filebeat_master.xlsx'

    check_input_file_existence(input_excel)
    filebeat_config(input_excel, master_excel)

    os.remove(input_excel)
    print("Processing complete.")


if __name__ == '__main__':
    main()

    


    